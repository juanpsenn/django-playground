from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import Alignment

def export_excel_report(title, header, data):
    # Create an instance of the Workbook class
    wb = Workbook()

    # Add a sheet to the workbook
    ws = wb.active
    ws.title = title

    # Write some data to the sheet
    ws.append(header)
    for row in data:
        ws.append(row)
    
    # Pepare final row
    ws.merge_cells(f'A{len(data)+1}:B{len(data)+1}')
    ws[f'A{len(data)+1}'].alignment = Alignment(horizontal="right", vertical="center")

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    # 0110111000111
    # ^
    # Seek to the beginning of the file
    output.seek(0)

    return output
